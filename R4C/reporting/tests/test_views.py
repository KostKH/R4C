import datetime as dt
from collections import defaultdict
from io import BytesIO

import openpyxl
from django.test import Client, TestCase
from django.urls import reverse

from robots.models import Robot


class ReportingViewsTests(TestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.guest_client = Client()
        cls.valid_data = [
            {
                'model': 'R2',
                'version': 'D2',
                'created': '2023-10-02 00:00:00',
            },
            {
                'model': 'R2',
                'version': 'D2',
                'created': '2023-10-08 23:59:59',
            },
            {
                'model': 'R2',
                'version': 'G7',
                'created': '2023-10-08 23:59:59',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-08 12:40:59',
            },
        ]
        cls.invalid_data = [
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-01 23:59:59',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-09 00:00:00',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2022-10-08 00:00:00',
            },
            {
                'model': 'I5',
                'version': 'K8',
                'created': '2022-10-08 00:00:00',
            },
        ]
        cls.test_data = cls.invalid_data + cls.valid_data

        for item in cls.test_data:
            Robot.objects.create(
                model=str(item['model']),
                version=str(item['version']),
                created=(dt.datetime.fromisoformat(item['created'])
                         .replace(tzinfo=dt.timezone.utc)),
                serial=f"{item['model']}-{item['version']}",
            )

    def test_pages_use_correct_template(self):
        """URL-адрес использует соответствующий шаблон."""
        templates_pages_names = [
            ('reporting/index.html', reverse('index')),
        ]
        for template, reverse_name in templates_pages_names:
            with self.subTest(reverse_name=reverse_name):
                response = self.guest_client.get(reverse_name)
                self.assertTemplateUsed(response, template)

    def test_index_shows_correct_context(self):
        """В шаблон index передан правильный контекст."""
        response = self.guest_client.get(reverse('index'))
        week = response.context.get('week')
        year = response.context.get('year')
        expected_week = dt.datetime.now().isocalendar().week
        expected_year = dt.datetime.now().year
        self.assertEqual(week, expected_week)
        self.assertEqual(year, expected_year)

    def test_production_report_shows_correct_content(self):
        """View-функция production_report отдает правильный контент."""
        year = 2023
        week = 40
        response = self.guest_client.get(
            reverse('production_report', args=[year, week]))
        file = BytesIO(response.content)
        workbook = openpyxl.load_workbook(file, read_only=True)
        expected_content_type = ('application/vnd.openxmlformats-'
                                 'officedocument.spreadsheetml.sheet')
        expected_content_disposition = f'filename=report_{week}_week.xlsx'
        expected_stat = defaultdict(lambda: defaultdict(int))
        for item in self.valid_data:
            expected_stat[item['model']][item['version']] += 1
        expected_sheetnames = ['total'] + sorted(expected_stat.keys())

        self.assertEqual(response._headers.get('content-type')[1],
                         expected_content_type)
        self.assertEqual(response._headers.get('content-disposition')[1],
                         expected_content_disposition)
        self.assertEqual(workbook.sheetnames, expected_sheetnames)

        total_cell = workbook.get_sheet_by_name('total').cell(2, 2)
        self.assertEqual(total_cell.value, len(self.valid_data))

        for model in expected_stat.keys():
            exp_version_count = len(expected_stat[model].keys())
            exp_model_total = sum(expected_stat[model].values())
            model_sheet = workbook.get_sheet_by_name(model)
            model_total_cell = model_sheet.cell(model_sheet.max_row, 3)
            with self.subTest(model=model):
                self.assertEqual(model_total_cell.value, exp_model_total)
                self.assertEqual(model_sheet.max_row - 2, exp_version_count)
        file.close()
