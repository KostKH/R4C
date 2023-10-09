import datetime as dt
from collections import defaultdict
from io import BytesIO

import openpyxl
from django.test import TestCase

from reporting.reports import ProductionReport
from robots.models import Robot


class ReportingReportsTests(TestCase):

    @classmethod
    def setUpClass(cls):
        super().setUpClass()
        cls.valid_data = [
            {
                'model': 'R2',
                'version': 'D2',
                'created': '2023-10-02 00:00:00',
            },
            {
                'model': 'R2',
                'version': 'D2',
                'created': '2023-10-08 23:59:59',
            },
            {
                'model': 'R2',
                'version': 'G7',
                'created': '2023-10-08 23:59:59',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-08 12:40:59',
            },
        ]
        cls.invalid_data = [
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-01 23:59:59',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2023-10-09 00:00:00',
            },
            {
                'model': 'S5',
                'version': 'K8',
                'created': '2022-10-08 00:00:00',
            },
            {
                'model': 'I5',
                'version': 'K8',
                'created': '2022-10-08 00:00:00',
            },
        ]
        cls.test_data = cls.invalid_data + cls.valid_data

        for item in cls.test_data:
            Robot.objects.create(
                model=str(item['model']),
                version=str(item['version']),
                created=(dt.datetime.fromisoformat(item['created'])
                         .replace(tzinfo=dt.timezone.utc)),
                serial=f"{item['model']}-{item['version']}",
            )

    def test_production_report_collects_correct_statistics(self):
        """Экземляр класса ProductionReport формирует
        корректный словарь со статистикой."""
        year = 2023
        week = 40
        expected_stat = defaultdict(lambda: defaultdict(int))
        for item in self.valid_data:
            expected_stat[item['model']][item['version']] += 1

        report = ProductionReport(week, year)
        self.assertEqual(report.total_robots, len(self.valid_data))
        for model in report.statistics.keys():
            for version in report.statistics[model].keys():
                with self.subTest(robot=(model, version)):
                    self.assertEqual(
                        report.statistics[model][version],
                        expected_stat[model][version]
                    )

    def test_write_report_method_saves_correct_file(self):
        """Экземляр класса ProductionReport сохраняет
        корректный словарь со статистикой."""
        year = 2023
        week = 40
        file = BytesIO()
        report = ProductionReport(week, year)
        report.write_report(file)

        workbook = openpyxl.load_workbook(file, read_only=True)
        expected_sheetnames = (
            ['total'] + sorted(report.statistics.keys())
        )
        self.assertEqual(workbook.sheetnames, expected_sheetnames)
        total_cell = workbook.get_sheet_by_name('total').cell(2, 2)
        self.assertEqual(total_cell.value, report.total_robots)

        for model in report.statistics.keys():
            exp_model_total = sum(report.statistics[model].values())
            exp_version_count = len(report.statistics[model].keys())
            model_sheet = workbook.get_sheet_by_name(model)
            model_total_cell = model_sheet.cell(model_sheet.max_row, 3)
            with self.subTest(model=model):
                self.assertEqual(model_total_cell.value, exp_model_total)
                self.assertEqual(model_sheet.max_row - 2, exp_version_count)
        file.close()
