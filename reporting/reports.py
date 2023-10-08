import datetime as dt
from collections import defaultdict
from typing import BinaryIO

import openpyxl

from robots.models import Robot


class ProductionReport():
    """Класс для формирования отчета по производству роботов за
    заданную неделю заданного года."""

    def __init__(self, week: int, year: int) -> None:
        self._week = week
        self._year = year
        self._get_statistics()

    def _get_statistics(self):
        """Метод запрашивает в БД список роботов, произведенных на неделе,
        и в год, указанные в параметрах конструктора класса, после чего
        преобразует полученные данные в словарь моделей со статистикой по
        версиям. Словарь сохраняется в атрибутах класса."""
        timezone = dt.timezone.utc
        date_start = (dt.datetime.fromisocalendar(self._year, self._week, 1)
                      .replace(tzinfo=timezone))
        date_end = (dt.datetime.fromisocalendar(self._year, self._week, 7)
                    .replace(tzinfo=timezone)
                    + dt.timedelta(seconds=86399, microseconds=999999))
        robots = (Robot.objects.filter(created__range=(date_start, date_end))
                  .values('model', 'version'))
        self.statistics = defaultdict(lambda: defaultdict(int))
        self.total_robots = 0
        for robot in robots:
            self.statistics[robot['model']][robot['version']] += 1
            self.total_robots += 1

    def write_report(self, output: BinaryIO) -> None:
        """Метод формирует Excel-отчет на основе сформированной статистики
        производства (берется из атрибута `statistics`) и сохраняет его
        в файл / file-like object."""
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'total'

        worksheet.append(
            ('Неделя', 'Производство всего (все модели)', 'Дата отчета')
        )
        worksheet.append((self._week, self.total_robots, dt.datetime.now()))
        self._make_cells_bold(worksheet, rows=(1, 1), columns=(1, 3))
        self._adjust_column_width(worksheet, row=1, columns=(1, 2))
        self._adjust_column_width(worksheet, row=2, columns=(3, 3))

        for model in sorted(self.statistics.keys()):
            worksheet = workbook.create_sheet(title=model)
            worksheet.append(('Модель', 'Версия', 'Количество за неделю'))
            self._make_cells_bold(worksheet, rows=(1, 1), columns=(1, 3))
            self._adjust_column_width(worksheet, row=1, columns=(1, 3))

            for version in sorted(self.statistics[model].keys()):
                worksheet.append(
                    (model, version, self.statistics[model][version])
                )

            worksheet.append(
                ('Всего', '', sum(self.statistics[model].values()))
            )
            self._make_cells_bold(
                worksheet,
                rows=(worksheet.max_row, worksheet.max_row),
                columns=(1, worksheet.max_column),
            )
        workbook.save(output)
        workbook.close()

    @staticmethod
    def _make_cells_bold(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        rows: list[int] | tuple[int],
        columns: list[int] | tuple[int],
    ) -> None:
        """Вспомогательный метод для выделения диапазона ячеек на
        листе Excel-книги жирным шрифотом."""
        bold_font = openpyxl.styles.Font(bold=True)
        for row in range(rows[0], rows[1] + 1):
            for column in range(columns[0], columns[1] + 1):
                worksheet.cell(row, column).font = bold_font

    @staticmethod
    def _adjust_column_width(
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        row: int,
        columns: list[int] | tuple[int],
    ) -> None:
        """Вспомогательный метод для изменения ширины ячеек у заданного
        диапазона на листе Excel-книги."""
        for column in range(columns[0], columns[1] + 1):
            column_letter = worksheet.cell(row, column).column_letter
            column_width = len(str(worksheet.cell(row, column).value)) + 3
            worksheet.column_dimensions[column_letter].width = column_width
